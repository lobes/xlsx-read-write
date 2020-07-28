import pandas as pd
import openpyxl as xl

fileName = "Process_Keywords.xlsx"
keywordDataFrame = pd.read_excel(fileName,engine="openpyxl")

# Get columns (each as a series). : means all rows and 0 means first column.
keywordsSeries = keywordDataFrame.iloc[:,0]
paliSeries = keywordDataFrame.iloc[:,1]

# Split keywords into separate words, with leading and trailing spaces trimmed
keywordListSplit = []
for cell in keywordsSeries:
    if isinstance(cell,str):  # Check for nan (empty cells)
        keywordList = cell.split(',')
        for word in keywordList:
            word = word.strip()
            if len(word) > 0:  # watch for empty strings
                keywordListSplit.append(word)

# Remove duplicates by making the list a dict
dictKeywords = dict.fromkeys(keywordListSplit)
keywordListSplitSorted = sorted(list(dictKeywords),key=str.lower)   # key=str.lower makes it case in-sensitive
# pandas ExcelWriter is needed to append
# Get value at first cell
#cellA1 = series1[0]
# Split comma delimited string into a list
#listSplitA1 = cellA1.split(',')
#print(cellA1)
print(keywordListSplitSorted)

# Split pali words into separate words, with leading and trailing spaces trimmed
paliListSplit = []
for cell in paliSeries:
    if isinstance(cell,str):  # Check for nan (empty cells)
        paliList = cell.split(',')
        for word in paliList:
            word = word.strip()
            if len(word) > 0:  # watch for empty strings
                paliListSplit.append(word)

# Remove duplicates by making the list a dict
dictPali = dict.fromkeys(paliListSplit)
paliListSplitSorted = sorted(list(dictPali), key=str.lower)   # key=str.lower makes it case-insensitive
print(paliListSplitSorted)

# Append the 2 dict objects (in case there is an english word = a pali word)
dictBoth = dict(dictKeywords)   # or dictKeywords.copy()
dictBoth.update(dictPali)

# Make a sorted list of the combined keywords-and-pali
bothListsSorted = sorted(list(dictBoth), key=str.lower)   # key=str.lower makes it case-insensitive
print(bothListsSorted)

wbkName = fileName
wbk = xl.load_workbook(wbkName)
wks = wbk['Topics for Book']

# Write the keywords out
rowIndex = 2  # row 2 will have first value
colIndex = 3  # 3rd column (C)
for word in keywordListSplitSorted:
    wks.cell(row=rowIndex, column=colIndex).value = word
    rowIndex +=1

# Write the pali out
rowIndex = 2  # row 2 will have first value
colIndex = 4  # 4th column (D)
for word in paliListSplitSorted:
    wks.cell(row=rowIndex, column=colIndex).value = word
    rowIndex +=1

# Write the combined list out
rowIndex = 2  # row 2 will have first value
colIndex = 5  # 5th column (E)
for word in bothListsSorted:
    wks.cell(row=rowIndex, column=colIndex).value = word
    rowIndex +=1

wbk.save(wbkName)
wbk.close
